"""
Export Service - PDF and Excel report generation
"""

import io
from typing import List, Dict, Any
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


def export_sales_to_pdf(sales_data: List[Dict], filename: str = None) -> bytes:
    """
    Export sales data to PDF.
    
    Args:
        sales_data: List of sale dictionaries
        filename: Optional filename
    
    Returns:
        PDF file as bytes
    """
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=1  # Center
        )
        title = Paragraph("Savdo Hisoboti", title_style)
        story.append(title)
        story.append(Spacer(1, 0.2*inch))
        
        # Date
        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        date_text = Paragraph(f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style)
        story.append(date_text)
        story.append(Spacer(1, 0.3*inch))
        
        # Summary
        total_amount = sum(sale.get('total_amount', 0) for sale in sales_data)
        total_items = sum(sale.get('item_count', 0) for sale in sales_data)
        
        summary_data = [
            ['Jami tranzaksiyalar:', str(len(sales_data))],
            ['Jami mahsulotlar:', str(total_items)],
            ['Jami summa:', f"{total_amount:,.0f} so'm"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Sales table
        if sales_data:
            table_data = [['ID', 'Sana', 'Mijoz', 'Mahsulotlar', 'Summa']]
            
            for sale in sales_data[:50]:  # Limit to 50 for PDF
                table_data.append([
                    str(sale.get('id', '')),
                    sale.get('created_at', '')[:10] if sale.get('created_at') else '',
                    sale.get('customer_name', 'N/A'),
                    str(sale.get('item_count', 0)),
                    f"{sale.get('total_amount', 0):,.0f}"
                ])
            
            sales_table = Table(table_data, colWidths=[0.5*inch, 1*inch, 1.5*inch, 1*inch, 1.5*inch])
            sales_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(sales_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        raise Exception(f"PDF yaratishda xatolik: {str(e)}")


def export_sales_to_excel(sales_data: List[Dict], filename: str = None) -> bytes:
    """
    Export sales data to Excel.
    
    Args:
        sales_data: List of sale dictionaries
        filename: Optional filename
    
    Returns:
        Excel file as bytes
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savdo Hisoboti"
        
        # Header
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A1'] = "Savdo Hisoboti"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A2:E2')
        
        # Summary
        total_amount = sum(sale.get('total_amount', 0) for sale in sales_data)
        ws['A4'] = "Jami tranzaksiyalar:"
        ws['B4'] = len(sales_data)
        ws['A5'] = "Jami summa:"
        ws['B5'] = f"{total_amount:,.0f} so'm"
        
        # Table headers
        headers = ['ID', 'Sana', 'Mijoz', 'Mahsulotlar', 'Summa (so\'m)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for row_idx, sale in enumerate(sales_data, start=8):
            ws.cell(row=row_idx, column=1, value=sale.get('id', ''))
            ws.cell(row=row_idx, column=2, value=sale.get('created_at', '')[:10] if sale.get('created_at') else '')
            ws.cell(row=row_idx, column=3, value=sale.get('customer_name', 'N/A'))
            ws.cell(row=row_idx, column=4, value=sale.get('item_count', 0))
            ws.cell(row=row_idx, column=5, value=sale.get('total_amount', 0))
        
        # Adjust column widths
        column_widths = [10, 15, 25, 15, 20]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        raise Exception(f"Excel yaratishda xatolik: {str(e)}")


def export_products_to_excel(products_data: List[Dict]) -> bytes:
    """Export products to Excel."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mahsulotlar"
        
        # Header
        ws['A1'] = "Mahsulotlar Ro'yxati"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')
        
        # Table headers
        headers = ['ID', 'Nomi', 'Narx', 'Xarajat', 'Omborda', 'Barcode', 'Kategoriya']
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, product in enumerate(products_data, start=4):
            ws.cell(row=row_idx, column=1, value=product.get('id', ''))
            ws.cell(row=row_idx, column=2, value=product.get('name', ''))
            ws.cell(row=row_idx, column=3, value=product.get('price', 0))
            ws.cell(row=row_idx, column=4, value=product.get('cost_price', 0))
            ws.cell(row=row_idx, column=5, value=product.get('stock_quantity', 0))
            ws.cell(row=row_idx, column=6, value=product.get('barcode', ''))
            ws.cell(row=row_idx, column=7, value=product.get('category_name', ''))
        
        # Adjust widths
        widths = [8, 30, 12, 12, 12, 15, 20]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error exporting products: {e}")
        raise Exception(f"Mahsulotlarni eksport qilishda xatolik: {str(e)}")








